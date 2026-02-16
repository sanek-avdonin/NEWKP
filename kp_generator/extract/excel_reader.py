from __future__ import annotations

from decimal import Decimal, InvalidOperation
from typing import List, Tuple

from openpyxl import load_workbook

from kp_generator.models import Item  # абсолютный импорт


def _norm(s: str) -> str:
    return "".join(ch for ch in s.lower().strip() if ch.isalnum() or ch in [" ", "_"]).replace("  ", " ")


REQUIRED = {
    "name": ["наименование", "товар", "услуга"],
    "qty": ["количество", "кол-во", "колво", "qty"],
    "unit": ["ед", "едизмерения", "ед измерения", "ед.", "ед.изм", "unit"],
    "price": ["цена", "цена за единицу", "стоимость", "price"],
    "amount": ["сумма", "итого", "стоимость всего", "amount"],
}


def _to_decimal(v) -> Decimal:
    if v is None:
        raise InvalidOperation()
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    s = str(v).replace(" ", "").replace(",", ".")
    return Decimal(s)


def read_items_from_excel(path: str) -> Tuple[List[Item], str]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    header_row = None
    col_map = {}

    for r in range(1, min(ws.max_row, 80) + 1):
        values = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 50) + 1)]
        normed = [_norm(str(v)) if v is not None else "" for v in values]

        def find_col(keys):
            for idx, cell in enumerate(normed, start=1):
                for k in keys:
                    if k in cell and cell != "":
                        return idx
            return None

        name_c = find_col(REQUIRED["name"])
        qty_c = find_col(REQUIRED["qty"])
        unit_c = find_col(REQUIRED["unit"])
        price_c = find_col(REQUIRED["price"])
        amount_c = find_col(REQUIRED["amount"])

        if all([name_c, qty_c, unit_c, price_c]):
            header_row = r
            col_map = {"name": name_c, "qty": qty_c, "unit": unit_c, "price": price_c, "amount": amount_c}
            break

    if header_row is None:
        raise ValueError("Не удалось найти заголовки таблицы в Excel. Проверьте названия колонок.")

    items: List[Item] = []

    for r in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(r, col_map["name"]).value
        if name is None or str(name).strip() == "":
            if len(items) > 0:
                break
            continue

        qty = _to_decimal(ws.cell(r, col_map["qty"]).value)
        unit = str(ws.cell(r, col_map["unit"]).value or "").strip()
        price = _to_decimal(ws.cell(r, col_map["price"]).value)

        amount_cell = None
        if col_map.get("amount"):
            amount_cell = ws.cell(r, col_map["amount"]).value

        if amount_cell is None or str(amount_cell).strip() == "":
            amount = (qty * price).quantize(Decimal("0.01"))
        else:
            amount = _to_decimal(amount_cell).quantize(Decimal("0.01"))

        items.append(Item(
            name=str(name).strip(),
            qty=qty,
            unit=unit,
            price=price.quantize(Decimal("0.01")),
            amount=amount
        ))

    if not items:
        raise ValueError("Таблица найдена, но строки товаров не извлечены.")

    return items, wb.active.title
