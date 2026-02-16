from __future__ import annotations

from decimal import Decimal
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from kp_generator.models import CompanyProfile, Item  # абсолютный импорт


PLACEHOLDERS = {
    "{{COMPANY_NAME}}": "name",
    "{{INN}}": "inn",
    "{{ADDRESS}}": "address",
    "{{PHONE}}": "phone",
    "{{CEO}}": "ceo",
}


def _replace_placeholders(ws: Worksheet, company: CompanyProfile) -> None:
    company_map = {
        "name": company.name,
        "inn": company.inn,
        "address": company.address,
        "phone": company.phone,
        "ceo": company.ceo,
    }
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "{{" in cell.value and "}}" in cell.value:
                v = cell.value
                for ph, key in PLACEHOLDERS.items():
                    v = v.replace(ph, str(company_map.get(key, "")))
                cell.value = v


def _norm(s: str) -> str:
    return "".join(ch for ch in s.lower().strip() if ch.isalnum() or ch in [" ", "_"]).replace("  ", " ")


def _find_table(ws: Worksheet) -> Tuple[int, Dict[str, int]]:
    req = {
        "name": ["наименование"],
        "qty": ["количество", "кол-во", "колво"],
        "unit": ["ед", "ед.", "едизм", "ед.изм", "ед измерения"],
        "price": ["цена"],
        "amount": ["сумма", "итого"],
    }

    for r in range(1, min(ws.max_row, 120) + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 60) + 1)]
        normed = [_norm(str(v)) if v is not None else "" for v in row_vals]

        def find_col(keys):
            for idx, cell in enumerate(normed, start=1):
                for k in keys:
                    if k in cell and cell != "":
                        return idx
            return None

        name_c = find_col(req["name"])
        qty_c = find_col(req["qty"])
        unit_c = find_col(req["unit"])
        price_c = find_col(req["price"])
        amount_c = find_col(req["amount"])

        if all([name_c, qty_c, unit_c, price_c]):
            return r, {"name": name_c, "qty": qty_c, "unit": unit_c, "price": price_c, "amount": amount_c}

    raise ValueError("Не удалось найти таблицу в шаблоне Excel по заголовкам.")


def _write_items(ws: Worksheet, header_row: int, col_map: Dict[str, int], items: List[Item]) -> int:
    start_row = header_row + 1

    r = start_row
    while r <= ws.max_row:
        v = ws.cell(r, col_map["name"]).value
        if v is None or str(v).strip() == "":
            break
        for key, col in col_map.items():
            ws.cell(r, col).value = None
        r += 1

    for i, it in enumerate(items):
        rr = start_row + i
        ws.cell(rr, col_map["name"]).value = it.name
        ws.cell(rr, col_map["qty"]).value = float(it.qty)
        ws.cell(rr, col_map["unit"]).value = it.unit
        ws.cell(rr, col_map["price"]).value = float(it.price)
        if col_map.get("amount"):
            ws.cell(rr, col_map["amount"]).value = float(it.amount)

    return start_row + len(items)


def _update_totals(ws: Worksheet, after_items_row: int, col_map: Dict[str, int], items: List[Item]) -> None:
    total = sum((it.amount for it in items), Decimal("0.00"))
    amount_col = col_map.get("amount") or col_map["price"]

    for r in range(after_items_row, min(ws.max_row, after_items_row + 60)):
        left = ws.cell(r, col_map["name"]).value
        if not isinstance(left, str):
            continue
        t = left.strip().lower()
        if t in ["итого", "всего", "итого:", "всего:"]:
            ws.cell(r, amount_col).value = float(total)
            break


def create_default_template(items: List[Item], company: CompanyProfile) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "КП"

    ws["A1"] = "Коммерческое предложение"
    ws["A3"] = "Компания: {{COMPANY_NAME}}"
    ws["A4"] = "ИНН: {{INN}}"
    ws["A5"] = "Адрес: {{ADDRESS}}"
    ws["A6"] = "Телефон: {{PHONE}}"
    ws["A7"] = "Руководитель: {{CEO}}"

    header_row = 10
    headers = ["Наименование", "Количество", "Ед. измерения", "Цена", "Сумма"]
    for i, h in enumerate(headers, start=1):
        ws.cell(header_row, i).value = h

    col_map = {"name": 1, "qty": 2, "unit": 3, "price": 4, "amount": 5}
    after_row = _write_items(ws, header_row, col_map, items)
    ws.cell(after_row + 1, 1).value = "Итого"
    ws.cell(after_row + 1, 5).value = float(sum((it.amount for it in items), Decimal("0.00")))

    _replace_placeholders(ws, company)
    return wb


def render_kp(
    template_path: Optional[str],
    company: CompanyProfile,
    items: List[Item],
    output_path: str
) -> None:
    if template_path:
        wb = load_workbook(template_path)
        ws = wb.active
        _replace_placeholders(ws, company)
        header_row, col_map = _find_table(ws)
        after_row = _write_items(ws, header_row, col_map, items)
        _update_totals(ws, after_row, col_map, items)
        wb.save(output_path)
    else:
        wb = create_default_template(items, company)
        wb.save(output_path)
